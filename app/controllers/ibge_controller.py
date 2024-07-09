import requests
import time

import os
from bs4 import BeautifulSoup
import pandas as pd
import shutil
from openpyxl import load_workbook
from datetime import datetime
from app.models import IPCA_IBGE, Desemprego_IBGE, CAGED_IBGE, db
from sqlalchemy.exc import IntegrityError


lista_url = [
    "https://apisidra.ibge.gov.br/values/t/7060/n1/all/v/63/p/first%206/c315/all/d/v63%202",
    "https://apisidra.ibge.gov.br/values/t/7060/n1/all/v/63/p/202007,202008,202009,202010,202011,202012/c315/all/d/v63%202",
    "https://apisidra.ibge.gov.br/values/t/7060/n1/all/v/63/p/202101,202102,202103,202104,202105,202106/c315/all/d/v63%202",
    "https://apisidra.ibge.gov.br/values/t/7060/n1/all/v/63/p/202107,202108,202109,202110,202111,202112/c315/all/d/v63%202",
    "https://apisidra.ibge.gov.br/values/t/7060/n1/all/v/63/p/202201,202202,202203,202204,202205,202206/c315/all/d/v63%202",
    "https://apisidra.ibge.gov.br/values/t/7060/n1/all/v/63/p/202207,202208,202209,202210,202211,202212/c315/all/d/v63%202",
    "https://apisidra.ibge.gov.br/values/t/7060/n1/all/v/63/p/202301,202302,202303,202304,202305,202306/c315/all/d/v63%202",
    "https://apisidra.ibge.gov.br/values/t/7060/n1/all/v/63/p/202307,202308,202309,202310,202311,202312/c315/all/d/v63%202",
    "https://apisidra.ibge.gov.br/values/t/7060/n1/all/v/63/p/last%205/c315/all/d/v63%202"
]

# def fetch_ipca_data():
#     all_data = []
#     for url in lista_url:
#         response = requests.get(url)
#         if response.status_code == 200:
#             all_data.extend(response.json())
#         else:
#             raise Exception(f"Erro na requisição IPCA: {response.status_code}")
#         time.sleep(10)  # Pausa de 10 segundos entre as requisições

#     # Filtrar dados para remover linhas indesejadas
#     filtered_data = [item for item in all_data if item['D4N'] != 'Geral, grupo, subgrupo, item e subitem']

#     # Extrair opções únicas para filtros e ordená-las
#     variable_options = sorted(list(set(item['D3C'] for item in filtered_data)))
#     unit_options = sorted(list(set(item['V'] for item in filtered_data)))
#     value_options = sorted(list(set(item['D4N'] for item in filtered_data)))

#     return filtered_data, variable_options, unit_options, value_options

def fetch_ipca_data():
    all_data = []
    for url in lista_url:
        response = requests.get(url)
        if response.status_code == 200:
            all_data.extend(response.json())
        else:
            raise Exception(f"Erro na requisição IPCA: {response.status_code}")
        time.sleep(10)  # Pausa de 10 segundos entre as requisições

    # Filtrar dados para remover linhas indesejadas
    filtered_data = [item for item in all_data if item['D4N'] != 'Geral, grupo, subgrupo, item e subitem']

    return filtered_data


def fetch_unemployment_data():
    url = "https://apisidra.ibge.gov.br/values/t/6381/n1/all/v/4099/p/all/d/v4099%201"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        
        # Filtrar dados para incluir apenas a partir de 201901
        filtered_data = [item for item in data if item['D3C'] >= '201901']
        
        return filtered_data
    else:
        raise Exception(f"Erro na requisição Desemprego: {response.status_code}")
    

# def download_caged_data():
#     url = "http://pdet.mte.gov.br/novo-caged"
#     response = requests.get(url)
#     soup = BeautifulSoup(response.text, 'html.parser')

#     links = soup.find_all("a", text="3. Tabelas.xlsx")
#     if links:
#         file_url = links[0]["href"]
#         full_file_url = "http://pdet.mte.gov.br" + file_url

#         response = requests.get(full_file_url)
#         if response.status_code == 200:
#             desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
#             download_folder = os.path.join(desktop_path, "IBGE-Postos de Trabalho")
#             os.makedirs(download_folder, exist_ok=True)
#             temp_file_path = os.path.join(download_folder, "temp_download.xlsx")

#             with open(temp_file_path, "wb") as file:
#                 file.write(response.content)

#             # Carregar o arquivo Excel usando pandas
#             xls = pd.ExcelFile(temp_file_path)

#             # Verificar se a aba "Tabela 8.1" existe
#             if "Tabela 8.1" in xls.sheet_names:
#                 df = pd.read_excel(xls, sheet_name="Tabela 8.1")

#                 # Corrigir os nomes das colunas removendo "Unnamed" e ajustando a formatação
#                 df.columns = [col if 'Unnamed' not in col else '' for col in df.columns]

#                 # Salvar a aba específica em um novo arquivo Excel
#                 final_file_path = os.path.join(download_folder, "Tabela_8_1.xlsx")
#                 with pd.ExcelWriter(final_file_path, engine='openpyxl') as writer:
#                     df.to_excel(writer, sheet_name="Tabela 8.1", index=False)

#                 # Fechar o arquivo Excel para liberar o recurso
#                 xls.close()

#                 # Mesclar as células no arquivo final
#                 wb = load_workbook(final_file_path)
#                 ws = wb["Tabela 8.1"]

#                 # Mesclar E5:H5
#                 ws.merge_cells('E5:H5')
#                 # Mesclar B5:B6
#                 ws.merge_cells('B5:B6')
#                 # Mesclar C5:C6
#                 ws.merge_cells('C5:C6')
#                 # Mesclar D5:D6
#                 ws.merge_cells('D5:D6')

#                 # Lógica para mesclar células a partir de I5 até que não haja mais conteúdo
#                 start_col = 9  # Coluna I
#                 row = 5

#                 while True:
#                     if ws.cell(row=row, column=start_col).value:
#                         end_col = start_col + 4  # Mescla 5 colunas à direita
#                         ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
#                         start_col = end_col + 1  # Avança para a próxima célula a verificar
#                     else:
#                         break

#                 wb.save(final_file_path)

#                 # Remover o arquivo temporário
#                 os.remove(temp_file_path)

#                 return final_file_path
#         else:
#             raise Exception("Erro ao baixar o arquivo")
#     else:
#         raise Exception("Elemento com a nome 'Tabela.xlsx' não encontrado")

def download_caged_data():
    url = "http://pdet.mte.gov.br/novo-caged"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')

    links = soup.find_all("a", text="3. Tabelas.xlsx")
    if links:
        file_url = links[0]["href"]
        full_file_url = "http://pdet.mte.gov.br" + file_url

        response = requests.get(full_file_url)
        if response.status_code == 200:
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            download_folder = os.path.join(desktop_path, "IBGE-Postos de Trabalho")
            os.makedirs(download_folder, exist_ok=True)
            temp_file_path = os.path.join(download_folder, "temp_download.xlsx")

            with open(temp_file_path, "wb") as file:
                file.write(response.content)

            # Carregar o arquivo Excel usando pandas
            xls = pd.ExcelFile(temp_file_path)

            # Verificar se a aba "Tabela 8.1" existe
            if "Tabela 8.1" in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name="Tabela 8.1", skiprows=4)  # Excluir as primeiras 4 linhas

                # Remover a primeira coluna (índice 0 em pandas)
                df = df.iloc[:, 1:]

                # Salvar o DataFrame como CSV
                final_csv_path = os.path.join(download_folder, "resultado.csv")
                df.to_csv(final_csv_path, index=False)

                # Fechar o arquivo Excel para liberar o recurso
                xls.close()

                # Remover o arquivo temporário
                os.remove(temp_file_path)

                return final_csv_path
        else:
            raise Exception("Erro ao baixar o arquivo")
    else:
        raise Exception("Elemento com a nome 'Tabela.xlsx' não encontrado")

    
def save_ipca_ibge_data(app):
    with app.app_context():
        IPCA_IBGE.query.delete()
        db.session.commit()
        
        datas = fetch_ipca_data()
        for item in datas:
            new_entry = IPCA_IBGE(
                periodo=item['D3C'],
                valor=item['V'],
                itens=item['D4N']
            )
            db.session.add(new_entry)
            print(f"Dados salvos: {new_entry}")
            db.session.commit()

def save_desemprego_ibge_data(app):
    with app.app_context():
        data = fetch_unemployment_data()
        for item in data:
            # Verifique se o valor é um número
            try:
                valor = float(item['V'])
            except ValueError:
                print(f"Valor inválido para o período {item['D3C']}: {item['V']}")
                continue

            if not Desemprego_IBGE.query.filter_by(data=item['D3C']).first():
                new_entry = Desemprego_IBGE(
                    valor=valor,
                    data=item['D3C']
                )
                db.session.add(new_entry)
                db.session.commit()
                print(f"Dados salvos: {new_entry}")
            else:
                print(f"Dados para o período {item['D3C']} já existem.")

def save_caged_data_to_db(app):
    # Baixar e processar os dados
    csv_path = download_caged_data()
    
    # Inicializar o contexto do aplicativo
    with app.app_context():
        # Carregar o CSV
        CAGED_IBGE.query.delete()
        db.session.commit()

        df = pd.read_csv(csv_path)
        df.columns = df.iloc[0].tolist()
        df = df[1:]
        df.rename(
            columns={
                df.columns[0]: 'UF',
                df.columns[1]: 'Código do Município',
                df.columns[2]: 'Município'
            }, inplace=True
        )

        col_aux = {}
        for nome in df.columns.unique().tolist():
            col_aux[nome] = 1

        col = []
        for column in df.columns:
            for key, value in col_aux.items():
                if key == column:
                    col.append(f"{column}_{value}")
                    col_aux[key] += 1
        df.columns = col
        df_auxiliar = pd.DataFrame(df)
        df_temp = df_auxiliar.iloc[:, 0:7]

        from datetime import datetime
        from dateutil.relativedelta import relativedelta

        data = datetime(2020, 1, 1)
        iteracao = 1
        while len(df_auxiliar.columns) > 12:
            if iteracao == 1:
                df_temp['Variação Relativa (%)'] = None
                df_temp['data_referencia'] = data
                df_auxiliar.drop(df_auxiliar.iloc[:, 3:7], axis=1, inplace=True)
            else:
                df_temp = df_auxiliar.iloc[:, 0:8]
                df_temp['data_referencia'] = data
                df_auxiliar.drop(df_auxiliar.iloc[:, 3:8], axis=1, inplace=True)
            # df_temp.to_csv(data.strftime('%m_%Y') + '.csv', index=False)
            data = data + relativedelta(months=1)
            
            list_aux = df_temp.columns.to_list()
            df_temp = df_temp.replace(['---', 'NaN', 'nan', pd.NaT, None], {list_aux[3]: 0, list_aux[4]: 0, list_aux[5]: 0, list_aux[6]: 0, list_aux[7]: 0.0})
            df_temp['Código do Município_1'] = pd.to_numeric(df_temp['Código do Município_1'], errors= 'coerce').fillna(0).astype(int)
            df_temp[list_aux[3]] = pd.to_numeric(df_temp[list_aux[3]], errors='coerce').fillna(0).astype(int)
            df_temp[list_aux[4]] = pd.to_numeric(df_temp[list_aux[4]], errors='coerce').fillna(0).astype(int)
            df_temp[list_aux[5]] = pd.to_numeric(df_temp[list_aux[5]], errors='coerce').fillna(0).astype(int)
            df_temp[list_aux[6]] = pd.to_numeric(df_temp[list_aux[6]], errors='coerce').fillna(0).astype(int)
            df_temp[list_aux[7]] = pd.to_numeric(df_temp[list_aux[7]], errors='coerce').fillna(0.0).astype(float)
            
            # Salvar os dados no banco de dados
            for index, row in df_temp.iterrows():
                caged_record = CAGED_IBGE(
                    UF=row['Município_1'],
                    Cod_Municipio=row[f'Código do Município_1'],
                    Municipio=row['Município_2'],
                    Mes=row['data_referencia'].strftime('%m_%Y'),
                    Estoque=row.get(list_aux[3], None),
                    Admissoes=row.get(list_aux[4], None),
                    Desligamentos=row.get(list_aux[5], None),
                    Saldos=row.get(list_aux[6], None),
                    Variacao=row.get(list_aux[7], None)
                )
                db.session.add(caged_record)
                print(f"Dados salvos: {caged_record.UF}, {caged_record.Municipio}, {caged_record.Mes}")
            
            db.session.commit()
            iteracao += 1
    

        